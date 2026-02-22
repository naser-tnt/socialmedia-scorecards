import os
import io
import csv
import zipfile
import shutil
import difflib
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path
import subprocess
import tempfile

import streamlit as st
import openpyxl

# Force page configuration to be wide for better preview
st.set_page_config(page_title="SocMedia Scorecards", layout="wide")

# Try to import playwright, if not available try to install it (useful for Streamlit Cloud)
try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    try:
        subprocess.run(["pip", "install", "playwright"], check=True)
        subprocess.run(["playwright", "install", "chromium"], check=True)
        from playwright.sync_api import sync_playwright
        PLAYWRIGHT_AVAILABLE = True
    except Exception:
        PLAYWRIGHT_AVAILABLE = False

EXCLUDED_PLACES = {"opi orders"}
EXCLUDED_STATUSES = {"cancelled", "rejected by place"}
DAY_NAMES = ["SUN", "MON", "TUES", "WED", "THUR", "FRI", "SAT"]

def normalize(name):
    """Strip emojis, normalize whitespace/case, replace & with 'and'."""
    if not name:
        return ""
    cleaned = "".join(c for c in name if ord(c) < 128)
    cleaned = cleaned.replace("&", "and")
    cleaned = " ".join(cleaned.split())
    return cleaned.strip().lower()

def parse_date(date_str):
    """Parse CSV date like '21 Feb 2026 11:57 pm'."""
    try:
        return datetime.strptime(date_str.strip(), "%d %b %Y %I:%M %p")
    except ValueError:
        return None

def day_of_week_sunday_start(dt):
    """Return 0=Sun, 1=Mon, ..., 6=Sat."""
    return (dt.weekday() + 1) % 7

def get_week_sunday(dt):
    """Return the Sunday that starts the week containing dt."""
    dow = day_of_week_sunday_start(dt)
    return (dt - timedelta(days=dow)).replace(hour=0, minute=0, second=0, microsecond=0)

@st.cache_data
def load_orders(file_content):
    """Load CSV orders, return list of dicts with parsed dates."""
    orders = []
    # Use io.StringIO to parse the uploaded bytes
    content = file_content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    for row in reader:
        place = row.get("Place", "").strip()
        status = row.get("Status", "").strip()
        date = parse_date(row.get("Date", ""))
        if not place or not date:
            continue
        if normalize(place) in EXCLUDED_PLACES:
            continue
        if status.lower() in EXCLUDED_STATUSES:
            continue
        orders.append({"place": place, "date": date, "norm": normalize(place)})
    return orders

@st.cache_data
def load_scorecard(file_content, filename):
    """Load scorecard data, supports both Excel (.xlsx) and CSV format, return dict of {norm_name: data}."""
    restaurants = {}
    
    if filename.lower().endswith(".csv"):
        # Handle CSV scorecard
        content = file_content.decode("utf-8-sig")
        reader = list(csv.reader(io.StringIO(content)))
        
        # In the provided CSV, header lines take up to row index 3
        # due to newlines wrapped in quotes. Data starts at row index 4.
        start_idx = 4
        for row in reader[start_idx:]:
            if not row or len(row) < 12:
                continue
            name = row[0]
            if not name or name.strip().lower() == "sum":
                continue
                
            norm = normalize(name)
            
            def val(col_val):
                cv = col_val.strip().upper()
                if cv == 'TRUE': return True
                if cv == 'FALSE': return False
                if cv == 'NA': return 'NA'
                return False
                
            # Indexes: 1=TipnTag, 2=IG, 3=FB, 4=Google, 5-11=Sun-Sat
            restaurants[norm] = {
                "display_name": name.strip(),
                "tip_tag": val(row[1]),
                "ig": val(row[2]),
                "fb": val(row[3]),
                "google": val(row[4]),
                "stories": [val(row[c]) for c in range(5, 12)],
            }
            true_count = sum(1 for c in range(2, 12) if val(row[c]) is True)
            restaurants[norm]["score"] = int((true_count / 10.0) * 100)
            
    else:
        # Handle Excel scorecard
        wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
        ws = wb[wb.sheetnames[0]]

        for row_idx in range(5, ws.max_row + 1):
            name = ws.cell(row=row_idx, column=1).value
            if not name or name.strip().lower() == "sum":
                continue
            norm = normalize(name)

            def val_xl(col):
                v = ws.cell(row=row_idx, column=col).value
                if v is True: return True
                if v is False: return False
                if isinstance(v, str) and v.strip().upper() == "NA": return "NA"
                return False

            restaurants[norm] = {
                "display_name": name.strip(),
                "tip_tag": val_xl(2),
                "ig": val_xl(3),
                "fb": val_xl(4),
                "google": val_xl(5),
                "stories": [val_xl(c) for c in range(6, 13)],  # F-L = Sun-Sat
            }
            
            true_count = sum(
                1 for c in range(3, 13)
                if ws.cell(row=row_idx, column=c).value is True
            )
            restaurants[norm]["score"] = int((true_count / 10.0) * 100)

        wb.close()
        
    return restaurants

def build_name_map(order_norms, scorecard_norms):
    """Build mapping from CSV normalized names → Excel normalized names using fuzzy matching."""
    mapping = {}
    for csv_n in order_norms:
        if csv_n in scorecard_norms:
            mapping[csv_n] = csv_n
            continue
            
        # Try direct substring matching first
        matched = False
        for sc_n in scorecard_norms:
            if sc_n in csv_n or csv_n in sc_n:
                mapping[csv_n] = sc_n
                matched = True
                break
                
        if not matched:
            # Fuzzy match as a fallback
            matches = difflib.get_close_matches(csv_n, scorecard_norms, n=1, cutoff=0.6)
            if matches:
                mapping[csv_n] = matches[0]
                
    return mapping

def count_orders_by_day(orders, name_map, week_start):
    """Count orders per restaurant per day-of-week for the given week."""
    week_end = week_start + timedelta(days=7)
    counts = defaultdict(lambda: [0] * 7)

    for o in orders:
        if o["date"] < week_start or o["date"] >= week_end:
            continue
        sc_name = name_map.get(o["norm"])
        if sc_name:
            dow = day_of_week_sunday_start(o["date"])
            counts[sc_name][dow] += 1
    return counts

def determine_week(orders):
    """Find the week (Sun-Sat) that contains most orders."""
    if not orders:
        now = datetime.now()
        ws = get_week_sunday(now)
        return ws, now.strftime("%B"), now.year, ((ws.day - 1) // 7) + 1

    week_counts = defaultdict(int)
    for o in orders:
        ws = get_week_sunday(o["date"])
        week_counts[ws] += 1

    week_start = max(week_counts, key=week_counts.get)
    month_name = week_start.strftime("%B")
    year = week_start.year
    week_num = ((week_start.day - 1) // 7) + 1
    return week_start, month_name, year, week_num

def generate_html(restaurant_data, daily_orders, month_name, year, week_num):
    """Generate the scorecard HTML for one restaurant strictly sized at 1000x1200."""
    name = restaurant_data["display_name"].upper()
    stories = restaurant_data["stories"]
    ig = restaurant_data["ig"]
    fb = restaurant_data["fb"]
    google = restaurant_data["google"]
    score = restaurant_data["score"]
    
    # max_orders scaling ensuring min is standard across
    max_orders = max(daily_orders) if max(daily_orders) > 0 else 1

    bars_html = ""
    for i, count in enumerate(daily_orders):
        posted = stories[i] is True
        color = "#28a745" if posted else "#dc3545" # More vibrant standard green/red
        if count > 0:
            height = max(30, int((count / max_orders) * 260))
            bars_html += f'''
            <div class="bar-col">
                <span class="bar-val" style="color:#111">{count}</span>
                <div class="bar" style="height:{height}px;background:{color}; border-radius: 6px 6px 0 0;"></div>
            </div>'''
        else:
            bars_html += f'''<div class="bar-col">
                <span class="bar-val" style="color:transparent">0</span>
                <div class="bar-empty"></div>
            </div>'''

    check_svg = '<svg width="36" height="36" viewBox="0 0 24 24" fill="none" stroke="#28a745" stroke-width="3.5" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"></polyline></svg>'
    cross_svg = '<svg width="36" height="36" viewBox="0 0 24 24" fill="none" stroke="#dc3545" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><line x1="18" y1="6" x2="6" y2="18"></line><line x1="6" y1="6" x2="18" y2="18"></line></svg>'

    story_cells = ""
    for s in stories:
        if s is True:
            story_cells += f'<td style="vertical-align: middle;">{check_svg}</td>'
        else:
            story_cells += f'<td style="vertical-align: middle;">{cross_svg}</td>'

    def link_icon(val):
        if val is True:
            return check_svg
        return cross_svg

    if score >= 70:
        score_color = "#28a745"
    elif score >= 50:
        score_color = "#ff9800"
    else:
        score_color = "#dc3545"

    import os
    svg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "B n B.svg")
    try:
        with open(svg_path, "r", encoding="utf-8") as f:
            svg_content = f.read()
    except Exception:
        svg_content = '<div style="font-size:24px;font-weight:900;color:#dc3545">biteME | tatbeeqi</div>'

    html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        * {{ margin:0; padding:0; box-sizing:border-box; font-family: 'Outfit', sans-serif; }}
        body {{ background: transparent; display: flex; justify-content: center; align-items: center; min-height: 100vh; margin: 0; }}
        
        .card {{ 
            width: 1000px;
            height: 1200px;
            padding: 50px 60px; 
            background: #fff; 
            border: 1px solid #fff; /* Removed container border for cleaner look */
            display: flex;
            flex-direction: column;
            justify-content: space-between;
        }}
        
        .header-container {{ display: flex; flex-direction: column; }}

        .header {{ display:flex; justify-content:space-between; align-items:center; margin-bottom: 20px; }}
        .header-left h2 {{ font-size: 38px; font-weight:900; line-height:1.2; color:#111; }}
        .header-left h2 span {{ font-weight:500; font-size: 28px; color:#333; }}
        
        .rest-title {{ font-size: 26px; font-weight:800; margin: 10px 0 0; color:#111; text-transform: uppercase; letter-spacing: 0.5px; }}
        
        .chart-section {{ flex: 1; display: flex; flex-direction: column; justify-content: flex-end; margin: 50px 0; }}
        .chart-title {{ text-align:center; font-weight:900; font-size: 26px; margin-bottom: 40px; color:#111; }}
        
        .chart-container {{ position: relative; flex: 1; min-height: 250px; border-bottom: 2px solid #eaeaea; }}
        
        /* Background grid lines */
        .grid-lines {{ position: absolute; inset: 0; display: flex; flex-direction: column; justify-content: space-between; z-index: 0; }}
        .grid-line {{ width: 100%; border-top: 1px solid #f8f8f8; }}
        
        .chart {{ display:flex; align-items:flex-end; justify-content:space-around; height:100%; position: absolute; inset:0; z-index: 1; padding:0 20px; }}
        
        .bar-col {{ display:flex; flex-direction:column; align-items:center; width: 80px; height: 100%; justify-content:flex-end; }}
        .bar {{ width: 70px; }}
        .bar-val {{ font-size: 24px; font-weight:700; margin-bottom: 10px; color: #111; }}
        .bar-empty {{ height:0; }}
        
        .table-container {{ margin-top: 20px; border: 2px solid #222; }}
        
        .section-hdr {{ background:#fff; color:#111; text-align:center; font-weight:800; font-size: 22px; padding: 18px 0; border-bottom: 2px solid #222; }}
        table {{ width:100%; border-collapse:collapse; background: #fff; table-layout: fixed; }}
        table td, table th {{ text-align:center; padding: 18px 10px; font-size: 22px; border-right: 2px solid #222; border-bottom: 2px solid #222; color: #111; }}
        table td:last-child, table th:last-child {{ border-right:none; }}
        table.no-bottom-border tr:last-child td {{ border-bottom:none; }}
        
        .day-hdr td {{ font-weight:600; font-size: 18px; text-transform:uppercase; color:#111; background:#fff; padding: 16px 10px; }}
        
        .score-val {{ font-size: 42px; font-weight:900; color:{score_color}; letter-spacing: -1px; line-height: 1; }}
    </style>
</head>
<body>
<div class="card" id="scorecard">
    <div class="header-container">
        <div class="header">
            <div class="header-left">
                <h2>Social Media Presence:<br><span>Score Card</span></h2>
            </div>
            <div class="logos" style="width: 260px; display: flex; justify-content: flex-end;">
                {svg_content}
            </div>
        </div>
        <div class="rest-title">{name}: {month_name} {year} &ndash; Week {week_num}</div>
    </div>

    <div class="chart-section">
        <div class="chart-title">Daily Orders (Excluding biteME)</div>
        <div class="chart-container">
            <div class="grid-lines">
                <div class="grid-line"></div>
                <div class="grid-line"></div>
                <div class="grid-line"></div>
                <div class="grid-line"></div>
                <div class="grid-line"></div>
            </div>
            <div class="chart">{bars_html}</div>
        </div>
    </div>

    <div class="table-container">
        <div class="section-hdr">Ordering Link Added in Story?</div>
        <table>
            <tr class="day-hdr">{''.join(f"<td>{d}</td>" for d in DAY_NAMES)}</tr>
            <tr>{story_cells}</tr>
        </table>
        
        <div class="section-hdr" style="border-top: none;">Permanent Link Added to ?</div>
        <table class="no-bottom-border">
            <tr class="day-hdr">
                <td>Instagram</td><td>Facebook</td><td>Google</td>
                <td style="font-size:16px;">Overall Score</td>
            </tr>
            <tr>
                <td>{link_icon(ig)}</td>
                <td>{link_icon(fb)}</td>
                <td>{link_icon(google)}</td>
                <td style="vertical-align: middle;"><span class="score-val">{score}%</span></td>
            </tr>
        </table>
    </div>
</div>
</body>
</html>'''
    return html

def html_to_png(html_path, png_path):
    """Convert HTML file to PNG using chromium headless browser directly to ensure consistent results"""
    abs_html = os.path.abspath(html_path)
    abs_png = os.path.abspath(png_path)
    
    # Using python-playwright API directly
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch()
            # Launch with viewport big enough to contain the 1000x1200 card safely with some body padding margins
            page = browser.new_page(viewport={"width": 1100, "height": 1300})
            page.goto(f"file://{abs_html}")
            # wait a small amount to make sure fonts are loaded
            page.wait_for_timeout(200)
            elem = page.locator(".card")
            elem.screenshot(path=abs_png)
            browser.close()
        return True
    except Exception as e:
        print(f"Error rendering {png_path}: {e}")
        return False

# --------------------------------------------------------------------------------
# STREAMLIT UI
# --------------------------------------------------------------------------------

st.title("📊 biteME Social Media Scorecard Automator")
st.markdown("Upload the **Daily Orders CSV** and the **Score Card Tracking (.csv or .xlsx)** to generate premium PNG scorecards for all restaurants.")

if not PLAYWRIGHT_AVAILABLE:
    st.warning("⚠️ Playwright is not installed. PNG generation works via standard python backend, make sure `pip install playwright && playwright install chromium` was run.")

col1, col2 = st.columns(2)

with col1:
    csv_file = st.file_uploader("Upload Bitesnbags Orders (.csv)", type=["csv"])

with col2:
    tracker_file = st.file_uploader("Upload Score Card tracking (.xlsx or .csv)", type=["xlsx", "csv"])

if csv_file and tracker_file:
    if st.button("🚀 Generate Scorecards", type="primary"):
        with st.spinner("Analyzing files and compiling data..."):
            
            # Load Data
            csv_bytes = csv_file.read()
            tracker_bytes = tracker_file.read()
            
            try:
                orders = load_orders(csv_bytes)
                scorecard = load_scorecard(tracker_bytes, tracker_file.name)
                
                # Determine dates
                week_start, month_name, year, week_num = determine_week(orders)
                
                # Match names
                order_norms = set(o["norm"] for o in orders)
                name_map = build_name_map(order_norms, set(scorecard.keys()))
                
                # Count
                daily_counts = count_orders_by_day(orders, name_map, week_start)
                
                # Sort exactly like what the user wanted: if they have orders, they appear;
                # Sort: restaurants with orders first, then alphabetically
                active_restaurants_in_csv = set(name_map.values())
                sorted_restaurants = sorted(
                    scorecard.items(),
                    key=lambda x: (
                        0 if x[0] in active_restaurants_in_csv and sum(daily_counts.get(x[0], [0])) > 0 else 1, 
                        x[1]["display_name"]
                    )
                )

                st.success(f"✅ Successfully compiled {len(sorted_restaurants)} restaurants for {month_name} Week {week_num}")
                
            except Exception as e:
                st.error(f"Error parsing data: {e}")
                st.stop()
                
        # Generate the images
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Temp dir for files
        temp_dir = tempfile.mkdtemp()
        html_dir = os.path.join(temp_dir, "html")
        png_dir = os.path.join(temp_dir, "scorecards")
        os.makedirs(html_dir, exist_ok=True)
        os.makedirs(png_dir, exist_ok=True)
        
        total = len(sorted_restaurants)
        generated_count = 0
        
        preview_htmls = []
        
        for i, (norm_name, data) in enumerate(sorted_restaurants):
            status_text.text(f"Generating scorecard for: {data['display_name']} ({i+1}/{total})")
            
            orders_per_day = daily_counts.get(norm_name, [0] * 7)
            safe = "".join(c for c in data["display_name"] if c.isalnum() or c in " -_").strip().replace(" ", "_")
            
            html = generate_html(data, orders_per_day, month_name, year, week_num)
            html_path = os.path.join(html_dir, f"{safe}.html")
            png_path = os.path.join(png_dir, f"{safe}.png")
            
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(html)
            
            if len(preview_htmls) < 4:
                preview_htmls.append((data['display_name'], html))
                
            if html_to_png(html_path, png_path):
                generated_count += 1
                
            progress_bar.progress((i + 1) / total)
            
        status_text.text("Compressing PNGs...")
        
        # Create Zip
        zip_path = os.path.join(temp_dir, f"scorecards_{month_name}_week{week_num}.zip")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(png_dir):
                for file in files:
                    zipf.write(os.path.join(root, file), arcname=file)
                    
        status_text.empty()
        progress_bar.empty()
        
        with open(zip_path, "rb") as f:
            zip_bytes = f.read()
            
        st.balloons()
        st.success(f"🎉 Generated {generated_count} scorecards successfully!")
        
        st.download_button(
            label="📦 Download All Scorecards (.zip)",
            data=zip_bytes,
            file_name=f"Social_Media_Scorecards_{month_name}_Week_{week_num}.zip",
            mime="application/zip",
            type="primary",
            use_container_width=True
        )
        
        st.markdown("### 👀 Previews")
        cols = st.columns(min(len(preview_htmls), 2))
        for i, (name, html) in enumerate(preview_htmls[:2]):
            with cols[i]:
                st.components.v1.html(html, height=750, scrolling=True)
        
        if len(preview_htmls) > 2:
            cols2 = st.columns(2)
            for i, (name, html) in enumerate(preview_htmls[2:4]):
                with cols2[i]:
                    st.components.v1.html(html, height=750, scrolling=True)

        # Cleanup temp
        shutil.rmtree(temp_dir, ignore_errors=True)
