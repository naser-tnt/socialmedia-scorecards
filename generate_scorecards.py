#!/usr/bin/env python3
"""
Social Media Scorecard Generator for bitesnbags
Reads order CSV + scorecard Excel to generate PNG scorecards per restaurant.
"""

import csv
import os
import re
import sys
import subprocess
import shutil
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ── Name Matching ──────────────────────────────────────────────────────────
# Maps normalized CSV names → normalized Excel names for tricky cases
MANUAL_NAME_MAP = {
    "shawerma 3a saj": "shawerma saj",
    "everybuddy nutrition supplements": "everybuddy",
    "pachi pizza and pasta": "pachi pizza",
    "azul pastry": "azul",
    "ikura japanese cuisine": "ikura",
    "the fit bar": "the fit bar jo",
    "sofia turkish restaurant": "sofia",
    "secrets cakes": "secrets cake",
    "flour and fire": "flour and fire",
}

EXCLUDED_PLACES = {"opi orders"}
EXCLUDED_STATUSES = {"cancelled", "rejected by place"}

DAY_NAMES = ["SUN", "MON", "TUES", "WED", "THUR", "FRI", "SAT"]


def normalize(name):
    """Strip emojis, normalize whitespace/case, replace & with 'and'."""
    if not name:
        return ""
    cleaned = "".join(c for c in name if ord(c) < 128)
    cleaned = cleaned.replace("&", "and")
    cleaned = " ".join(cleaned.split())
    return cleaned.strip().lower()


def parse_date(date_str):
    """Parse CSV date like '21 Feb 2026 11:57 pm'."""
    try:
        return datetime.strptime(date_str.strip(), "%d %b %Y %I:%M %p")
    except ValueError:
        return None


def day_of_week_sunday_start(dt):
    """Return 0=Sun, 1=Mon, ..., 6=Sat."""
    return (dt.weekday() + 1) % 7


def get_week_sunday(dt):
    """Return the Sunday that starts the week containing dt."""
    dow = day_of_week_sunday_start(dt)
    return (dt - timedelta(days=dow)).replace(hour=0, minute=0, second=0, microsecond=0)


# ── Load Data ──────────────────────────────────────────────────────────────

def load_orders(csv_path):
    """Load CSV orders, return list of dicts with parsed dates."""
    orders = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            place = row.get("Place", "").strip()
            status = row.get("Status", "").strip()
            date = parse_date(row.get("Date", ""))
            if not place or not date:
                continue
            if normalize(place) in EXCLUDED_PLACES:
                continue
            if status.lower() in EXCLUDED_STATUSES:
                continue
            orders.append({"place": place, "date": date, "norm": normalize(place)})
    return orders


def load_scorecard(xlsx_path):
    """Load Excel scorecard, return dict of {norm_name: data}."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    restaurants = {}
    for row_idx in range(5, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        if not name or name.strip().lower() == "sum":
            continue
        norm = normalize(name)

        def val(col):
            v = ws.cell(row=row_idx, column=col).value
            if v is True:
                return True
            if v is False:
                return False
            if isinstance(v, str) and v.strip().upper() == "NA":
                return "NA"
            return False

        restaurants[norm] = {
            "display_name": name.strip(),
            "tip_tag": val(2),
            "ig": val(3),
            "fb": val(4),
            "google": val(5),
            "stories": [val(c) for c in range(6, 13)],  # F-L = Sun-Sat
        }
        # Calculate score: count of TRUE in C-L / 10
        true_count = sum(
            1
            for c in range(3, 13)
            if ws.cell(row=row_idx, column=c).value is True
        )
        restaurants[norm]["score"] = int(true_count * 10)

    wb.close()
    return restaurants


# ── Match Names ────────────────────────────────────────────────────────────

def build_name_map(order_norms, scorecard_norms):
    """Build mapping from CSV normalized names → Excel normalized names."""
    mapping = {}
    for csv_n in order_norms:
        # Direct match
        if csv_n in scorecard_norms:
            mapping[csv_n] = csv_n
            continue
        # Manual map
        if csv_n in MANUAL_NAME_MAP and MANUAL_NAME_MAP[csv_n] in scorecard_norms:
            mapping[csv_n] = MANUAL_NAME_MAP[csv_n]
            continue
        # Contains match (CSV name contains Excel name or vice versa)
        for sc_n in scorecard_norms:
            if sc_n in csv_n or csv_n in sc_n:
                mapping[csv_n] = sc_n
                break
    return mapping


# ── Count Orders ───────────────────────────────────────────────────────────

def count_orders_by_day(orders, name_map, week_start):
    """Count orders per restaurant per day-of-week for the given week."""
    week_end = week_start + timedelta(days=7)
    counts = defaultdict(lambda: [0] * 7)

    for o in orders:
        if o["date"] < week_start or o["date"] >= week_end:
            continue
        sc_name = name_map.get(o["norm"])
        if sc_name:
            dow = day_of_week_sunday_start(o["date"])
            counts[sc_name][dow] += 1
    return counts


# ── Determine Week ─────────────────────────────────────────────────────────

def determine_week(orders):
    """Find the week (Sun-Sat) that contains most orders."""
    if not orders:
        now = datetime.now()
        ws = get_week_sunday(now)
        return ws, now.strftime("%B"), now.year, ((ws.day - 1) // 7) + 1

    week_counts = defaultdict(int)
    for o in orders:
        ws = get_week_sunday(o["date"])
        week_counts[ws] += 1

    week_start = max(week_counts, key=week_counts.get)
    month_name = week_start.strftime("%B")
    year = week_start.year
    week_num = ((week_start.day - 1) // 7) + 1
    return week_start, month_name, year, week_num


# ── HTML Generation ────────────────────────────────────────────────────────

def generate_html(restaurant_data, daily_orders, month_name, year, week_num):
    """Generate the scorecard HTML for one restaurant."""
    name = restaurant_data["display_name"].upper()
    stories = restaurant_data["stories"]
    ig = restaurant_data["ig"]
    fb = restaurant_data["fb"]
    google = restaurant_data["google"]
    score = restaurant_data["score"]
    max_orders = max(daily_orders) if max(daily_orders) > 0 else 1

    # Bar chart HTML
    bars_html = ""
    for i, count in enumerate(daily_orders):
        posted = stories[i] is True
        color = "#4CAF50" if posted else "#e53935"
        if count > 0:
            height = max(30, int((count / max_orders) * 150))
            bars_html += f'''
            <div class="bar-col">
                <span class="bar-val" style="color:{color}">{count}</span>
                <div class="bar" style="height:{height}px;background:{color};"></div>
            </div>'''
        else:
            bars_html += '<div class="bar-col"><div class="bar-empty"></div></div>'

    # Story row
    story_cells = ""
    for s in stories:
        if s is True:
            story_cells += '<td><span class="check">&#10003;</span></td>'
        else:
            story_cells += '<td><span class="cross">&#10007;</span></td>'

    # Link row
    def link_icon(val):
        if val is True:
            return '<span class="check">&#10003;</span>'
        return '<span class="cross">&#10007;</span>'

    score_color = "#4CAF50" if score >= 50 else "#e53935"

    html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family: Arial, Helvetica, sans-serif; background:#fff; }}
.card {{ width:460px; padding:20px 24px 24px; background:#fff; }}
.header {{ display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:6px; }}
.header-left h2 {{ font-size:14px; font-weight:800; line-height:1.3; }}
.header-left h2 span {{ font-weight:400; }}
.logos {{ display:flex; gap:8px; align-items:center; }}
.logo-box {{ background:#e53935; color:#fff; font-weight:900; font-size:13px;
    padding:4px 8px; border-radius:4px; font-family:Arial; }}
.logo-box2 {{ background:#e53935; color:#fff; font-weight:700; font-size:10px;
    padding:4px 8px; border-radius:4px; text-align:center; line-height:1.2; }}
.logo-box2 small {{ font-size:7px; display:block; font-weight:400; }}
.rest-title {{ font-size:13px; font-weight:800; margin:8px 0 14px; }}
.chart-title {{ text-align:center; font-weight:800; font-size:14px; margin-bottom:8px; }}
.chart {{ display:flex; align-items:flex-end; justify-content:space-around;
    height:190px; padding:10px 5px 0; border-bottom:2px solid #ccc; }}
.bar-col {{ display:flex; flex-direction:column; align-items:center;
    width:50px; justify-content:flex-end; }}
.bar {{ width:38px; border-radius:3px 3px 0 0; }}
.bar-val {{ font-size:12px; font-weight:700; margin-bottom:2px; }}
.bar-empty {{ height:0; }}
.section {{ margin-top:16px; border:1px solid #ddd; border-radius:4px; overflow:hidden; }}
.section-hdr {{ background:#333; color:#fff; text-align:center;
    font-weight:700; font-size:13px; padding:7px 0; }}
table {{ width:100%; border-collapse:collapse; }}
table td, table th {{ text-align:center; padding:6px 2px; font-size:12px;
    border-right:1px solid #eee; }}
table td:last-child, table th:last-child {{ border-right:none; }}
.day-hdr {{ font-weight:700; font-size:11px; border-bottom:1px solid #ddd; }}
.check {{ color:#4CAF50; font-size:22px; font-weight:700; }}
.cross {{ color:#e53935; font-size:22px; font-weight:700; }}
.link-section table td {{ padding:8px 4px; }}
.score-val {{ font-size:28px; font-weight:900; color:{score_color}; }}
.score-label {{ font-size:10px; color:#888; font-style:italic; }}
</style></head>
<body><div class="card">
  <div class="header">
    <div class="header-left">
      <h2>Social Media Presence:<br><span>Score Card</span></h2>
    </div>
    <div class="logos">
      <div class="logo-box">bite<span style="font-size:15px">ME</span></div>
      <div class="logo-box2"><small>POWERED BY</small>tabbeeqi</div>
    </div>
  </div>

  <div class="rest-title">{name}: {month_name} {year} &ndash; Week {week_num}</div>

  <div class="chart-title">Daily Orders (Excluding biteME)</div>
  <div class="chart">{bars_html}</div>

  <div class="section">
    <div class="section-hdr">Ordering Link Added in Story?</div>
    <table>
      <tr class="day-hdr">{''.join(f"<td>{d}</td>" for d in DAY_NAMES)}</tr>
      <tr>{story_cells}</tr>
    </table>
  </div>

  <div class="section link-section">
    <div class="section-hdr">Permanent Link Added to ?</div>
    <table>
      <tr class="day-hdr">
        <td>Instagram</td><td>Facebook</td><td>Google</td>
        <td colspan="1"><span class="score-label">Overall Score</span></td>
      </tr>
      <tr>
        <td>{link_icon(ig)}</td>
        <td>{link_icon(fb)}</td>
        <td>{link_icon(google)}</td>
        <td><span class="score-val">{score}%</span></td>
      </tr>
    </table>
  </div>
</div></body></html>'''
    return html


# ── PNG Conversion ─────────────────────────────────────────────────────────

def html_to_png(html_path, png_path):
    """Convert HTML file to PNG using Chrome headless or playwright."""
    abs_html = os.path.abspath(html_path)

    # Try Chrome headless
    chrome_paths = [
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        shutil.which("google-chrome"),
        shutil.which("chromium"),
    ]
    for chrome in chrome_paths:
        if chrome and os.path.exists(chrome):
            try:
                subprocess.run(
                    [chrome, "--headless=new", "--disable-gpu", "--no-sandbox",
                     f"--screenshot={os.path.abspath(png_path)}",
                     "--window-size=500,800", f"file://{abs_html}"],
                    capture_output=True, timeout=30,
                )
                if os.path.exists(png_path):
                    return True
            except Exception:
                continue

    # Try playwright
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page(viewport={"width": 500, "height": 800})
            page.goto(f"file://{abs_html}")
            page.wait_for_timeout(500)
            page.locator(".card").screenshot(path=os.path.abspath(png_path))
            browser.close()
        return True
    except Exception as e:
        print(f"  ⚠ PNG conversion failed: {e}")
        return False


# ── Main ───────────────────────────────────────────────────────────────────

def main():
    base = Path(__file__).parent
    csv_files = sorted(base.glob("*.csv"))
    xlsx_files = sorted(base.glob("Score Card*.xlsx"))

    if not csv_files:
        print("ERROR: No CSV file found in", base)
        sys.exit(1)
    if not xlsx_files:
        print("ERROR: No Score Card Excel file found in", base)
        sys.exit(1)

    csv_path = csv_files[0]
    xlsx_path = xlsx_files[0]
    print(f"CSV:   {csv_path.name}")
    print(f"Excel: {xlsx_path.name}")

    # Load data
    orders = load_orders(str(csv_path))
    scorecard = load_scorecard(str(xlsx_path))
    print(f"Loaded {len(orders)} orders, {len(scorecard)} restaurants in scorecard")

    # Determine week
    week_start, month_name, year, week_num = determine_week(orders)
    week_end = week_start + timedelta(days=6)
    print(f"Week {week_num}: {week_start.strftime('%b %d')} - {week_end.strftime('%b %d, %Y')}")

    # Build name mapping
    order_norms = set(o["norm"] for o in orders)
    name_map = build_name_map(order_norms, set(scorecard.keys()))
    matched = set(name_map.values())
    unmatched_csv = order_norms - set(name_map.keys())
    if unmatched_csv:
        print(f"Unmatched CSV places (no scorecard): {unmatched_csv}")

    # Count orders per day
    daily_counts = count_orders_by_day(orders, name_map, week_start)

    # Create output dirs
    html_dir = base / "output" / "html"
    png_dir = base / "output" / "png"
    html_dir.mkdir(parents=True, exist_ok=True)
    png_dir.mkdir(parents=True, exist_ok=True)

    # Sort: restaurants with orders first, then alphabetically
    sorted_restaurants = sorted(
        scorecard.items(),
        key=lambda x: (0 if x[0] in daily_counts else 1, x[1]["display_name"]),
    )

    generated = 0
    for norm_name, data in sorted_restaurants:
        orders_per_day = daily_counts.get(norm_name, [0] * 7)
        safe = re.sub(r"[^\w\s-]", "", data["display_name"]).strip().replace(" ", "_")

        html = generate_html(data, orders_per_day, month_name, year, week_num)
        html_path = html_dir / f"{safe}.html"
        png_path = png_dir / f"{safe}.png"

        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html)

        if html_to_png(str(html_path), str(png_path)):
            print(f"  ✓ {data['display_name']}")
        else:
            print(f"  ~ {data['display_name']} (HTML only)")
        generated += 1

    print(f"\nDone! Generated {generated} scorecards")
    print(f"  HTML: {html_dir}")
    print(f"  PNG:  {png_dir}")


if __name__ == "__main__":
    main()
